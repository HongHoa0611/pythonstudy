from tkinter import *
from tkinter import ttk
from MHDangNhap import man_hinh_dang_nhap
import json
from openpyxl import Workbook


root = Tk()
root.title("Quản Lý Sản Phẩm")
root.geometry("1000x400")
#menu chính (main menu)
menu_bar = Menu(root)
#nhúng menu vào màn hình chính
root.config(menu=menu_bar)

menu_file = Menu(menu_bar, tearoff=False)
menu_file.add_command(label="Đăng nhập", command=lambda: man_hinh_dang_nhap(root))
menu_file.add_command(label="Đăng xuất")  
menu_file.add_separator()
menu_file.add_command(label="Đóng chương trình")

menu_bar.add_cascade(label="Hệ Thống", menu=menu_file)
def doc_du_lieu_san_pham():
    global data_san_pham # dạng global
    try:
        with open("dataSanPham.json", "r", encoding="utf-8") as myfile:
            data_san_pham = json.load(myfile)
    except:
        data_san_pham = []


#gắn menu export:
def xuat_file_xlxs():
    
    wb = Workbook() #lúc tạo đã có 1 sheet default
    ws = wb.active #lấy sheet dang active
    #ws = wb.create_sheet("ANC")

    #gán cột tiêu đề
    ws["A1"] = "Mã"
    ws["B1"] = "Tên Hàng Hóa"
    ws["C1"] = "Loại"
    ws["D1"] = "Đơn giá"
    ws["E1"] = "Số Lượng"
    
    idong = 2
    for product in data_san_pham:
        ws["A" + str(idong)] = product["ma_sp"]
        ws["B" + str(idong)] = product["ten_sp"]
        ws["C" + str(idong)] = product["loai"]
        ws["D" + str(idong)] = product["don_gia"]
        ws["E" + str(idong)] = product["so_luong"]
        idong += 1
    

    wb.save("san_pham.xlsx")
mnu_export = Menu(menu_bar, tearoff=False)
mnu_export.add_command(label="Export excel file", command=xuat_file_xlxs)
menu_bar.add_cascade(label="Export", menu=mnu_export)

# Chia làm 3 frame: Danh sách (TreeView), thao tác (Button), chi tiết (nhiều widget bên trong...)
frame_danh_sach = ttk.Frame(root)
frame_tac_vu = ttk.Frame(root)
frame_chi_tiet = ttk.Frame(root)
frame_danh_sach.grid(row=0, column=0)
frame_tac_vu.grid(row=1, column=0)
frame_chi_tiet.grid(row=2, column=0)

# Vùng danh sách
san_pham_columns = ('ma_sp', 'ten_sp', 'don_gia', 'so_luong', 'loai')
tv = ttk.Treeview(frame_danh_sach, columns=san_pham_columns, show='headings', height=4)
tv.pack()

# Vùng tác vụ (gồm các button)
ttk.Button(frame_tac_vu, text="Thêm").pack()

doc_du_lieu_san_pham()
xuat_file_xlxs()

root.mainloop()