from tkinter import *
from tkinter import ttk
from MHDangNhap import man_hinh_dang_nhap
from tkinter.messagebox import showinfo, showerror, showwarning, askquestion
from openpyxl import Workbook

root = Tk()
root.title("X1 datalist")
root.geometry("1000x400")
#menu chính (main menu)
menu_bar = Menu(root)
#nhúng menu vào màn hình chính
root.config(menu=menu_bar)

menu_file = Menu(menu_bar, tearoff=False)
menu_file.add_command(label="log in", command=lambda: man_hinh_dang_nhap(root))
menu_file.add_command(label="log out")  
menu_file.add_separator()
menu_file.add_command(label="Exit")

menu_bar.add_cascade(label="Menu", menu=menu_file)
def doc_du_lieu_data():
    global data_list # dạng global
    try:
        with open("photo_list.json", "r", encoding="utf-8") as myfile:
            data_list = json.load(myfile)
    except:
        data_list = []


#gắn menu export:
def xuat_file_xlxs():
    wb = Workbook() #lúc tạo đã có 1 sheet default
    ws = wb.active #lấy sheet dang active
    #ws = wb.create_sheet("ANC")

    #gán cột tiêu đề
    ws["A1"] = "data_name"
    ws["B1"] = "data_collector"
    ws["C1"] = "specimen"
    ws["D1"] = "data_type"
    ws["E1"] = "image_size"
    
    idong = 2
    for product in data_san_pham:
        ws["A" + str(idong)] = product["data_name"]
        ws["B" + str(idong)] = product["data_collector"]
        ws["C" + str(idong)] = product["specimen"]
        ws["D" + str(idong)] = product["data_type"]
        ws["E" + str(idong)] = product["image_size"]
        idong += 1
    

    wb.save("data_list.xlsx")

#gắn menu export:
mnu_export = Menu(menu_bar, tearoff=False)
mnu_export.add_command(label="Export excel file", command=xuat_file_xlxs)
menu_bar.add_cascade(label="Export", menu=mnu_export)

# Chia làm 3 frame: Danh sách (TreeView), thao tác (Button), chi tiết (nhiều widget bên trong...)
frame_list = ttk.Frame(root)
frame_tac_vu = ttk.Frame(root)
frame_chi_tiet = ttk.Frame(root)
frame_list.grid(row=0, column=0)
frame_tac_vu.grid(row=1, column=0)
frame_chi_tiet.grid(row=2, column=0)

# Vùng danh sách
data_columns = ('data_id', 'data_name', 'specimen', 'data_type', 'image_size')
tv = ttk.Treeview(frame_list, columns=data_columns, show='headings', height=4)
tv.pack()

# Vùng tác vụ (gồm các button)
ttk.Button(frame_tac_vu, text="Add").pack()

doc_du_lieu_data()
xuat_file_xlxs()
root.mainloop()